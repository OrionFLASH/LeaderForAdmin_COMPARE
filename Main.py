import os
import json
import pandas as pd
import re
import logging
import traceback
# import sys
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# --- Параметры логирования ---
# LOG_LEVEL определяет глубину вывода в консоль (INFO или DEBUG)
LOG_LEVEL = logging.INFO

# === Константы путей и имён файлов ===
# Здесь задаются пути к папкам с исходниками, результатами и логами,
# а также имена входных/выходных файлов. При необходимости их можно
# поменять под свои каталоги.
SOURCE_DIR = "//Users//orionflash//Desktop//MyProject//LeaderForAdmin_skript//JSON"
TARGET_DIR = "//Users//orionflash//Desktop//MyProject//LeaderForAdmin_skript//XLSX"
LOG_DIR = "//Users//orionflash//Desktop//MyProject//LeaderForAdmin_skript//LOGS"
LOG_BASENAME = "LOG_INFO"
BEFORE_FILENAME = "leadersForAdmin_ALL_20250714-093911.json"
AFTER_FILENAME = "leadersForAdmin_ALL_20250725-135321.json"
RESULT_EXCEL = "LFA_COMPARE.xlsx"
# === Параметры справочников турниров и конкурсов ===
CATALOG_DIR = "//Users//orionflash//Desktop//MyProject//LeaderForAdmin_skript//CSV"
TOURNAMENT_SCHEDULE_CSV = "TOURNAMENT-SCHEDULE (PROM) 2025-07-24 v3.csv"
CONTEST_DATA_CSV = "CONTEST-DATA (PROM) 2025-07-24 v4.csv"


# --- Список турниров, которые будут включены в анализ ---
# Если список пустой, сравниваются все турниры из исходных файлов.
ALLOWED_TOURNAMENT_IDS = [
    "t_01_2025-2_08-2_6_2031",
    "t_01_2025-2_09-1_1_3071",
    "t_01_2025-2_14-1_1_3071",
    "t_01_2025-2_16-1_1_2031",
    "t_01_2025-2_05-1_1_3071",
    "t_01_2025-2_08-1_1_3071",
    "t_01_2025-2_15-1_2_2031",
    "t_01_2025-2_01-7_1_4001",
    "t_01_2025-1_07-5_5_4001",
    "t_01_2025-1_07-5_6_4001",
    "t_01_2025-1_07-5_7_4001",
    "t_01_2025-0_18-7_5_4001",
    "t_01_2025-1_04-1_1_4001"
]

# --- Фильтрация турниров в листах BEFORE и AFTER ---
# Если True, то на листы BEFORE и AFTER будут загружены только турниры из ALLOWED_TOURNAMENT_IDS
# Если ALLOWED_TOURNAMENT_IDS пустой, то грузятся все турниры независимо от этого параметра
# Если False, то всегда грузятся все турниры
FILTER_TOURNAMENTS_IN_BEFORE_AFTER = True

LOG_MESSAGES = {
    "LOGGER_SESSION_START": "\n-------- NEW LOG START AT {date} ({time}) -------\n",
    "LOGGER_ACTIVE_FILE": "Лог-файл активен (append): {path}",
    "DATAFRAME_EMPTY": "[{label}] DataFrame пустой.",
    "DATAFRAME_SHAPE": "[{label}] строк: {n_rows}, колонок: {n_cols}",
    "TOURNAMENT_ID_LIST": "[{label}] tournamentId всего: {count} -> {tids}",
    "TOURNAMENT_ID_PERSON_COUNT": "[{label}] tournamentId={tid}: людей={count}",
    "DATAFRAME_COLUMNS": "[{label}] Все поля: {cols}",
    "COMPARE_TOTAL_ROWS": "[COMPARE] Строк всего: {n_rows}",
    "COMPARE_COLUMN_COUNTS": "[COMPARE] {col}: {counts}",
    "PARSE_FLOAT_ERROR": "[parse_float] Ошибка преобразования '{val}' в float: {ex} | Context: {context}",
    "PARSE_INT_ERROR": "[parse_int] Ошибка преобразования '{val}' в int: {ex} | Context: {context}",
    "FLATTEN_LEADER_START": "Начата обработка лидера: employee={employee} для турнира {tournament_id}, файл {source_file}",
    "PROCESS_JSON_LOAD_ERROR": "Ошибка загрузки файла {filepath}: {ex}",
    "PROCESS_JSON_BAD_RECORD": "[process_json_file] Некорректная запись в турнире {tournament_key}: {record}",
    "PROCESS_JSON_EMPTY_LEADERS": "Турнир {tournament_id} из файла {filename}: leaders пуст, добавлена заглушка",
    "PROCESS_JSON_FLATTEN_LEADER_ERROR": "[flatten_leader] Ошибка обработки лидера в файле {filename} турнир {tournament_id} employee {employee}: {ex}",
    "PROCESS_JSON_RECORD_ERROR": "[process_json_file] Ошибка обработки записи в файле {filename}, турнир {tournament_key}: {ex}",
    "LOAD_JSON_NO_DATA": "Нет данных для экспорта из папки {folder}",
    "COMPARE_SHEET_START":    "[COMPARE] Построение листа сравнения...",
    "COMPARE_SHEET_FILTERED": "[COMPARE] После фильтрации по tournamentId: {count} строк.",
    "COMPARE_SHEET_FINAL":    "[COMPARE] После фильтрации строк без изменений: {count} строк.",
    "SMART_TABLE_EXPORT_START":   "[EXPORT] Экспорт листа {sheet} в Excel...",
    "SMART_TABLE_FORMATTED":      "[EXPORT] Отформатирован лист {sheet} (автоширина, автофильтр, freeze).",
    "SMART_TABLE_ERROR":          "[EXPORT][ERROR] Ошибка при экспорте/форматировании листа '{sheet}': {err}",
    "STATUS_COLOR_START":         "[EXPORT] Применяется цветовая раскраска к листу {sheet} по статусам.",
    "STATUS_COLOR_DONE":          "[EXPORT] Цветовая раскраска завершена для {sheet}.",
    "STATUS_LEGEND_ADD":          "[EXPORT] Добавлена легенда по статусам на лист {sheet}.",
    "FINAL_BUILD_START":        "=== [FINAL] Построение итоговой сводной таблицы ===",
    "FINAL_UNIQUE_EMPLOYEES":   "[FINAL] Уникальных сотрудников: {num_employees}",
    "FINAL_TOTAL_LOOPS":        "[FINAL] Всего итераций обработки: {loops}",
    "FINAL_TABLE_DONE":         "[FINAL] Итоговая таблица построена: {shape}",
    "FINAL_TOURN_STATUS":       "[{sheet}] tournamentId={tid} - распределение статусов:",
    "FINAL_TOURN_STATUS_ROW":   "[{sheet}]     {status}: {count}",
    "PLACE_BUILD_START":      "=== [{sheet}] Построение итоговой таблицы по placeInRating ===",
    "PLACE_UNIQUE_EMPLOYEES": "[{sheet}] Уникальных сотрудников: {num_employees}",
    "PLACE_TOURNAMENTS":      "[{sheet}] Турниров: {num_tournaments}",
    "PLACE_TABLE_DONE":       "[{sheet}] Итоговая таблица построена: {shape}",
    "PLACE_TOURN_STATUS":     "[{sheet}] tournamentId={tid} - распределение статусов:",
    "PLACE_TOURN_STATUS_ROW": "[{sheet}]     {status}: {count}",
    "STATUSES_ADD_START":   "[{sheet}] Добавление итоговых колонок по статусам: {statuses}",
    "STATUSES_ADDED_COL":   "[{sheet}] Статус '{status}': всего по таблице {total}",
    "STATUSES_ROW_DETAIL":  "[{sheet}] [{emp_info}] Итоги по статусам: {stat_counts}",
    "ADD_STATUSES_SUMMARY": "[ADD_STATUSES] Добавлены итоговые колонки: {columns}",
    "EXPORT_SHEET":         "[MAIN] Экспортирован лист {sheet} ({rows} строк).",
    "COND_FMT_START": "[{sheet}] Применяется условное форматирование для колонок с префиксами {prefixes}",
    "COND_FMT_COL":   "[{sheet}] Применено условное форматирование для колонки '{col}'",
    "COND_FMT_FINISH": "[{sheet}] Завершено условное форматирование для всех колонок с префиксами {prefixes}",
    "MAIN_BEFORE_READ": "[MAIN] Читаем {sheet}: {path}",
    "MAIN_BEFORE_LOADED": "[MAIN] Загружено {count} строк из {sheet}.",
    "MAIN_AFTER_READ": "[MAIN] Читаем {sheet}: {path}",
    "MAIN_AFTER_LOADED": "[MAIN] Загружено {count} строк из {sheet}.",
    "MAIN_TOURNAMENTS_INFO": "[MAIN] Турниров в {sheet_before}: {before_count}, в {sheet_after}: {after_count}",
    "MAIN_TOURNAMENTS_NEW": "[MAIN] Новые турниры (только в {sheet_after}): {count} -> {ids}",
    "MAIN_TOURNAMENTS_REMOVED": "[MAIN] Удалённые турниры (только в {sheet_before}): {count} -> {ids}",
    "MAIN_TOURNAMENTS_COMMON": "[MAIN] Общие турниры: {count} -> {ids}",
    "MAIN_COLUMNS_ALIGNED": "[MAIN] Приведены колонки {sheet_before} и {sheet_after} к единой структуре.",
    "MAIN_COMPARE_DONE": "[MAIN] Построен {sheet}: {count} строк.",
    "MAIN_FINAL_DONE": "[MAIN] Построен {sheet}: {shape}",
    "MAIN_FINAL_PLACE_DONE": "[MAIN] Построен {sheet}: {shape}",
    "MAIN_FINAL_STATUS_SUMMARY": "[MAIN] Добавлены сводные колонки по статусам в {sheet}.",
    "MAIN_FINAL_PLACE_STATUS_SUMMARY": "[MAIN] Добавлены сводные колонки по статусам в {sheet}.",
    "MAIN_FINAL_TOP3": "[MAIN] Добавлены колонки TOP1/2/3 и групп в {sheet}.",
    "MAIN_FINAL_PLACE_TOP3": "[MAIN] Добавлены колонки TOP1/2/3 в {sheet}.",
    "MAIN_EXCEL_EXPORT": "[MAIN] Все данные успешно выгружены в файл: {path}",
    "MAIN_STAT_COND_FMT": "[MAIN] Применено условное форматирование к stat_/grp_ в листе {sheet}.",
    "MAIN_COLORS_APPLIED": "[MAIN] Применена цветовая раскраска к {sheet}.",
    "MAIN_LEGEND_ADDED": "[MAIN] Добавлен лист с легендой статусов.",
    "MAIN_FINAL_SET_ACTIVE_SHEET_FAIL": "[MAIN] Не удалось установить лист {sheet} активным: {ex}",
    "MAIN_TOURNAMENT_DESCRIPTIONS_LOADED": "[MAIN] Загружено описаний турниров: {count}"
}

# Шаблон итоговой строки
SUMMARY_TEMPLATE_EXT = (
    "[SUMMARY] турниров: {tourn}; сотрудников: {emps}; изменений: {changes};\n"
    "Время: load_before: {t1:.2f}s; load_after: {t2:.2f}s; compare: {t3:.2f}s; final: {t4:.2f}s; export: {t5:.2f}s; total: {tt:.2f}s\n"
    "FINAL (распределение по статусам): {final_statuses}\n"
    "FINAL_PLACE (распределение по статусам): {final_place_statuses}\n"
    "FINAL (распределение по группам): {final_groups}\n"
)

# --- Статусы, при которых считаем, что изменений не произошло ---
# Если встречается один из этих кодов, строка считается без изменений
NOCHANGE_STATUSES = [
#    "", "Тот же индикатор", "Остался вне призеров", "Сохранил призовую позицию", "Индикатор пропал",
#    "Место пропало (BANK)", "Место пропало (TB)", "Место пропало (GOSB)",
#    "Такое же место (BANK)", "Такое же место (TB)", "Такое же место (GOSB)",
#    "Не участвовал", "Нет места"
]

# --- Основные колонки в исходных данных ---
# Эти поля всегда присутствуют и выводятся в итоговую таблицу в первую очередь
PRIORITY_COLS = [
    'SourceFile', 'tournamentId', 'employeeNumber', 'lastName', 'firstName',
    'terDivisionName', 'divisionRatings_BANK_groupId', 'divisionRatings_TB_groupId',
    'divisionRatings_GOSB_groupId', 'employeeStatus', 'businessBlock',
    'successValue', 'indicatorValue', 'divisionRatings_BANK_placeInRating',
    'divisionRatings_TB_placeInRating', 'divisionRatings_GOSB_placeInRating',
    'divisionRatings_BANK_ratingCategoryName', 'divisionRatings_TB_ratingCategoryName',
    'divisionRatings_GOSB_ratingCategoryName'
]
# Ключевые поля, по которым склеиваются данные "до" и "после"
COMPARE_KEYS = [
    'tournamentId',
    'employeeNumber',
    'lastName',
    'firstName'
]

# --- Карта закрепления (freeze panes) ---
freeze_map = {
        "FINAL":        "D2",  # Первая строка + 3 столбца
        "FINAL_PLACE":  "D2",
        "COMPARE":      "E2",  # Первая строка + 4 столбца
        "BEFORE":       "G2",  # Первая строка + 6 столбцов
        "AFTER":        "G2"
}

# Поля, которые сравниваются между выгрузками
COMPARE_FIELDS = [
    'SourceFile',
    'terDivisionName',
    'divisionRatings_TB_groupId',
    'divisionRatings_GOSB_groupId',
    'indicatorValue',
    'divisionRatings_BANK_placeInRating',
    'divisionRatings_TB_placeInRating',
    'divisionRatings_GOSB_placeInRating',
    'divisionRatings_BANK_ratingCategoryName',
    'divisionRatings_TB_ratingCategoryName',
    'divisionRatings_GOSB_ratingCategoryName'
]

COMPARE_EXPORT_COLUMNS = [
    'tournamentId', 'employeeNumber', 'lastName', 'firstName',
    'SourceFile',  # объединённое
    'terDivisionName',  # объединённое
    'divisionRatings_TB_groupId',  # объединённое
    'divisionRatings_GOSB_groupId',  # объединённое

    'BEFORE_indicatorValue', 'AFTER_indicatorValue', 'indicatorValue_Compare',

    # Объединенные статусы мест (лучший доступный уровень)
    'BEFORE_placeInRating_Best', 'AFTER_placeInRating_Best', 'placeInRating_Compare_Best', 'placeInRating_Level',
    
    # Объединенные статусы категорий (лучший доступный уровень)  
    'BEFORE_ratingCategoryName_Best', 'AFTER_ratingCategoryName_Best', 'ratingCategoryName_Compare_Best', 'ratingCategoryName_Level',

    # Исходные поля по уровням (для справки)
    'BEFORE_divisionRatings_BANK_placeInRating', 'AFTER_divisionRatings_BANK_placeInRating', 'divisionRatings_BANK_placeInRating_Compare',
    'BEFORE_divisionRatings_TB_placeInRating', 'AFTER_divisionRatings_TB_placeInRating', 'divisionRatings_TB_placeInRating_Compare',
    'BEFORE_divisionRatings_GOSB_placeInRating', 'AFTER_divisionRatings_GOSB_placeInRating', 'divisionRatings_GOSB_placeInRating_Compare',

    'BEFORE_divisionRatings_BANK_ratingCategoryName', 'AFTER_divisionRatings_BANK_ratingCategoryName', 'divisionRatings_BANK_ratingCategoryName_Compare',
    'BEFORE_divisionRatings_TB_ratingCategoryName', 'AFTER_divisionRatings_TB_ratingCategoryName', 'divisionRatings_TB_ratingCategoryName_Compare',
    'BEFORE_divisionRatings_GOSB_ratingCategoryName', 'AFTER_divisionRatings_GOSB_ratingCategoryName', 'divisionRatings_GOSB_ratingCategoryName_Compare',
]

# Поля, которые должны быть приведены к типу int
INT_FIELDS = [
    'divisionRatings_BANK_groupId',
    'divisionRatings_TB_groupId',
    'divisionRatings_GOSB_groupId',
    'divisionRatings_BANK_placeInRating',
    'divisionRatings_TB_placeInRating',
    'divisionRatings_GOSB_placeInRating'
]
# Поля, которые хранят вещественные значения
FLOAT_FIELDS = [
    'indicatorValue',
    'successValue',
]

# Все возможные статусы для итоговых колонок (FINAL)
FINAL_STATUS_LIST = [
    "Новый призёр", "Поднялся в рейтинге призеров", "Стал призёром", "Сохранил призовую позицию",
    "Снизил призовое место", "Лишился награды", "Удалённый призёр", "Остался вне призеров",
    "Новый участник без награды", "Удалённый участник без награды"
]

# Для FINAL_PLACE нужны только PLACE-статусы:
FINAL_PLACE_STATUS_LIST = [
    "Улучшил место (BANK)", "Улучшил место (TB)", "Улучшил место (GOSB)",
    "Хуже место (BANK)", "Хуже место (TB)", "Хуже место (GOSB)",
    "Новый с местом (BANK)", "Новый с местом (TB)", "Новый с местом (GOSB)",
    "Место пропало (BANK)", "Место пропало (TB)", "Место пропало (GOSB)",
    "Такое же место (BANK)", "Такое же место (TB)", "Такое же место (GOSB)"
]

STATUS_GROUPS = [
    ("Группа 1", ["Новый призёр", "Поднялся в рейтинге призеров", "Стал призёром"]),
    ("Группа 2", ["Сохранил призовую позицию"]),
    ("Группа 3", ["Снизил призовое место"]),
    ("Группа 4", ["Лишился награды", "Удалённый призёр"]),
    ("Группа 5", ["Остался вне призеров", "Новый участник без награды", "Удалённый участник без награды"]),
    ("Группа 6", ["Не участвовал"])
]

# --- Все цвета статусов здесь ---
STATUS_COLORS_DICT = {
    # ==== Новая шкала статусов по ratingCategoryName ====
    "Новый призёр":                 {"fill": "#00B050", "font": "#000000"},  # Ярко-зелёный, чёрный текст
    "Снизил призовое место":        {"fill": "#FFC7CE", "font": "#000000"},  # Светло-красный
    "Поднялся в рейтинге призеров": {"fill": "#C6EFCE", "font": "#000000"},  # Светло-зелёный
    "Сохранил призовую позицию":    {"fill": "#D9EAD3", "font": "#000000"},  # Светло-зелёный оттенок
    "Лишился награды":              {"fill": "#FF0000", "font": "#FFFFFF"},  # Красный, белый текст
    "Стал призёром":                {"fill": "#00B0F0", "font": "#000000"},  # Ярко-синий
    "Остался вне призеров":                {"fill": "#BFBFBF", "font": "#000000"},  # Серый
    "Удалённый участник без награды": {"fill": "#808080", "font": "#FFFFFF"},# Тёмно-серый, белый текст
    "Удалённый призёр":             {"fill": "#808080", "font": "#FFFFFF"},  # Тёмно-серый, белый текст
    "Новый участник без награды":   {"fill": "#E2EFDA", "font": "#000000"},  # Бледно-зелёный

    # ==== Прочие статусы ====
    "Индикатор упал":                  {"fill": "#FFC7CE", "font": "#000000"},
    "Индикатор вырос":                    {"fill": "#C6EFCE", "font": "#000000"},
    "Новый индикатор":              {"fill": "#E2EFDA", "font": "#000000"},
    "Индикатор пропал":             {"fill": "#383838", "font": "#FFFFFF"},  # Тёмно-серый, белый текст
    "Нет места":                    {"fill": "#F5F5F5", "font": "#C8C8C8"},
    "Нет призового значения":       {"fill": "#F5F5F5", "font": "#C8C8C8"},
    "Нет значения ранга":           {"fill": "#F5F5F5", "font": "#C8C8C8"},
    "Не участвовал":                {"fill": "#F5F5F5", "font": "#C8C8C8"},  # Почти белый, бледно-серый текст

    # ==== Статусы по PLACE рейтингу ====
    'Улучшил место (BANK)':                 {"fill": "#C6EFCE", "font": "#000000"},
    'Улучшил место (TB)':                   {"fill": "#C6EFCE", "font": "#000000"},
    'Улучшил место (GOSB)':                 {"fill": "#C6EFCE", "font": "#000000"},
    'Хуже место (BANK)':               {"fill": "#FFC7CE", "font": "#000000"},
    'Хуже место (TB)':                 {"fill": "#FFC7CE", "font": "#000000"},
    'Хуже место (GOSB)':               {"fill": "#FFC7CE", "font": "#000000"},
    'Новый с местом (BANK)':                {"fill": "#E2EFDA", "font": "#000000"},
    'Новый с местом (TB)':                  {"fill": "#E2EFDA", "font": "#000000"},
    'Новый с местом (GOSB)':                {"fill": "#E2EFDA", "font": "#000000"},
    'Место пропало (BANK)':             {"fill": "#383838", "font": "#FFFFFF"},
    'Место пропало (TB)':               {"fill": "#383838", "font": "#FFFFFF"},
    'Место пропало (GOSB)':             {"fill": "#383838", "font": "#FFFFFF"},
    'Такое же место (BANK)':          {"fill": "#BFBFBF", "font": "#000000"},
    'Такое же место (TB)':            {"fill": "#BFBFBF", "font": "#000000"},
    'Такое же место (GOSB)':          {"fill": "#BFBFBF", "font": "#000000"},
}

# --- Статусы для листа FINAL_PLACE (placeInRating) ---
ALL_STATUSES_PLACE = [
    "Улучшил место (BANK)", "Улучшил место (TB)", "Улучшил место (GOSB)",
    "Новый с местом (BANK)", "Новый с местом (TB)", "Новый с местом (GOSB)",
    "Такое же место (BANK)", "Такое же место (TB)", "Такое же место (GOSB)",
    "Нет места",
    "Нет значения ранга",
    "Не участвовал",
    "Хуже место (BANK)", "Хуже место (TB)", "Хуже место (GOSB)",
    "Место пропало (BANK)", "Место пропало (TB)", "Место пропало (GOSB)",
]

# --- Ключевые статусные колонки для сравнения, фильтрации, раскраски ---
COMPARE_STATUS_COLUMNS = [
    'indicatorValue_Compare',
    # Объединенные статусы (основные для анализа)
    'placeInRating_Compare_Best',
    'ratingCategoryName_Compare_Best',
    # Исходные статусы по уровням (для детального анализа)
    'divisionRatings_BANK_placeInRating_Compare',
    'divisionRatings_TB_placeInRating_Compare',
    'divisionRatings_GOSB_placeInRating_Compare',
    'divisionRatings_BANK_ratingCategoryName_Compare',
    'divisionRatings_TB_ratingCategoryName_Compare',
    'divisionRatings_GOSB_ratingCategoryName_Compare'
]

# --- Справочник по статусам (Excel-код: (рус, комментарий)) ---
STATUS_LEGEND_FULL = [
    # --- FINAL: Категории ---
    ("Новый призёр",                   "NewWinner",        "Участник впервые попал в призёры",                "#00B050", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 1"),
    ("Поднялся в рейтинге призеров",   "UpInWinners",      "Улучшил место среди призёров",                    "#C6EFCE", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 1"),
    ("Стал призёром",                  "BecameWinner",     "Был без награды, попал в призёры",                "#00B0F0", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 1"),
    ("Сохранил призовую позицию",      "KeptPlace",        "Остался на том же призовом месте",                "#D9EAD3", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 2"),
    ("Снизил призовое место",           "DownInWinners",    "Место стал хуже среди призёров",                  "#FFC7CE", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 3"),
    ("Лишился награды",                "LostAward",        "Был призёром, стал без награды",                  "#FF0000", "#FFFFFF", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 4"),
    ("Удалённый призёр",               "RemovedWinner",    "Был призёром, но исчез из списка",                "#808080", "#FFFFFF", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 4"),
    ("Остался вне призеров",           "StayedNoAward",    "Был вне призёров и остался вне",                  "#BFBFBF", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 5"),
    ("Новый участник без награды",     "NewNoAward",       "Появился впервые, но не стал призёром",           "#E2EFDA", "#000000", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 5"),
    ("Удалённый участник без награды", "RemovedNoAward",   "Был в прошлом, исчез, не был призёром",           "#808080", "#FFFFFF", "FINAL, divisionRatings_*_ratingCategoryName_Compare", "Группа 5"),
    ("Нет призового значения",         "NoPrizeRank",      "Был в выгрузках, но нет ранга/категории",         "#F5F5F5", "#C8C8C8", "FINAL", "-"),
    ("Не участвовал",                  "NotParticipated",  "Отсутствует в обеих выгрузках",                   "#F5F5F5", "#C8C8C8", "FINAL, FINAL_PLACE", "Группа 6"),

    # --- FINAL_PLACE: Места ---
    ("Улучшил место (BANK)",           "ImprovedPlace_BANK", "Место улучшилось по BANK",                       "#C6EFCE", "#000000", "FINAL_PLACE, divisionRatings_BANK_placeInRating_Compare", "-"),
    ("Улучшил место (TB)",             "ImprovedPlace_TB",   "Место улучшилось по TB",                         "#C6EFCE", "#000000", "FINAL_PLACE, divisionRatings_TB_placeInRating_Compare", "-"),
    ("Улучшил место (GOSB)",           "ImprovedPlace_GOSB", "Место улучшилось по GOSB",                       "#C6EFCE", "#000000", "FINAL_PLACE, divisionRatings_GOSB_placeInRating_Compare", "-"),
    ("Хуже место (BANK)",              "WorsePlace_BANK",    "Место стало хуже по BANK",                       "#FFC7CE", "#000000", "FINAL_PLACE, divisionRatings_BANK_placeInRating_Compare", "-"),
    ("Хуже место (TB)",                "WorsePlace_TB",      "Место стало хуже по TB",                         "#FFC7CE", "#000000", "FINAL_PLACE, divisionRatings_TB_placeInRating_Compare", "-"),
    ("Хуже место (GOSB)",              "WorsePlace_GOSB",    "Место стало хуже по GOSB",                       "#FFC7CE", "#000000", "FINAL_PLACE, divisionRatings_GOSB_placeInRating_Compare", "-"),
    ("Новый с местом (BANK)",          "NewWithPlace_BANK",  "Появилось место по BANK",                        "#E2EFDA", "#000000", "FINAL_PLACE, divisionRatings_BANK_placeInRating_Compare", "-"),
    ("Новый с местом (TB)",            "NewWithPlace_TB",    "Появилось место по TB",                          "#E2EFDA", "#000000", "FINAL_PLACE, divisionRatings_TB_placeInRating_Compare", "-"),
    ("Новый с местом (GOSB)",          "NewWithPlace_GOSB",  "Появилось место по GOSB",                        "#E2EFDA", "#000000", "FINAL_PLACE, divisionRatings_GOSB_placeInRating_Compare", "-"),
    ("Место пропало (BANK)",           "LostPlace_BANK",     "Место исчезло по BANK",                          "#383838", "#FFFFFF", "FINAL_PLACE, divisionRatings_BANK_placeInRating_Compare", "-"),
    ("Место пропало (TB)",             "LostPlace_TB",       "Место исчезло по TB",                            "#383838", "#FFFFFF", "FINAL_PLACE, divisionRatings_TB_placeInRating_Compare", "-"),
    ("Место пропало (GOSB)",           "LostPlace_GOSB",     "Место исчезло по GOSB",                          "#383838", "#FFFFFF", "FINAL_PLACE, divisionRatings_GOSB_placeInRating_Compare", "-"),
    ("Такое же место (BANK)",          "SamePlace_BANK",     "Место не изменилось по BANK",                    "#BFBFBF", "#000000", "FINAL_PLACE, divisionRatings_BANK_placeInRating_Compare", "-"),
    ("Такое же место (TB)",            "SamePlace_TB",       "Место не изменилось по TB",                      "#BFBFBF", "#000000", "FINAL_PLACE, divisionRatings_TB_placeInRating_Compare", "-"),
    ("Такое же место (GOSB)",          "SamePlace_GOSB",     "Место не изменилось по GOSB",                    "#BFBFBF", "#000000", "FINAL_PLACE, divisionRatings_GOSB_placeInRating_Compare", "-"),
    ("Нет места",                      "NoPlace",            "Место отсутствует (нет ранга)",                  "#F5F5F5", "#C8C8C8", "FINAL_PLACE, *_placeInRating_Compare", "-"),
    ("Нет значения ранга",             "NoRankValue",        "В выгрузке был, но нет числа места",             "#F5F5F5", "#C8C8C8", "FINAL_PLACE", "-"),

    # --- Индикаторы ---
    ("Индикатор вырос",                "IndicatorUp",        "Показатель вырос",                               "#C6EFCE", "#000000", "COMPARE, indicatorValue_Compare", "-"),
    ("Индикатор упал",                 "IndicatorDown",      "Показатель упал",                                "#FFC7CE", "#000000", "COMPARE, indicatorValue_Compare", "-"),
    ("Новый индикатор",                "NewIndicator",       "Появился показатель",                            "#E2EFDA", "#000000", "COMPARE, indicatorValue_Compare", "-"),
    ("Индикатор пропал",               "IndicatorLost",      "Показатель исчез",                               "#383838", "#FFFFFF", "COMPARE, indicatorValue_Compare", "-"),
    ("Тот же индикатор",               "SameIndicator",      "Показатель не изменился",                        "#BFBFBF", "#000000", "COMPARE, indicatorValue_Compare", "-"),
    ("",                               "NoChange",           "Нет изменений",                                  "#F5F5F5", "#C8C8C8", "COMPARE", "-"),
]


# --- Группы для FINAL и их описания для легенды и GRP_MAX ---
GROUP_DESC_DICT = {
    "Группа 1": "Поднялся в рейтинге призеров, Новый призёр, Стал призёром",
    "Группа 2": "Сохранил призовую позицию",
    "Группа 3": "Снизил призовое место",
    "Группа 4": "Лишился награды, Удалённый призёр",
    "Группа 5": "Остался вне призеров, Новый участник без награды, Удалённый участник без награды",
    "Группа 6": "Не участвовал",
}

# --- Названия листов для экспорта ---
SHEET_NAMES = {
    "before": "BEFORE",
    "after": "AFTER",
    "compare": "COMPARE",
    "final": "FINAL",
    "final_place": "FINAL_PLACE",
    "status_legend": "STATUS_LEGEND",
}

# --- Статусы сравнения для показателя indicatorValue ---
STATUS_INDICATOR = {
    "val_add":      "Новый индикатор",      # Значение появилось (было None/отсутствовало, стало не None)
    "val_remove":   "Индикатор пропал",  # Значение исчезло (было не None, стало None)
    "val_nochange": "Тот же индикатор",    # Значение не изменилось (равно до и после)
    "val_down":     "Индикатор упал",  # Значение уменьшилось (стало ниже)
    "val_up":       "Индикатор вырос"     # Значение увеличилось (стало выше)
}

# --- Статусы сравнения для места (placeInRating) BANK ---
STATUS_BANK_PLACE = {
    "val_add":      "Новый с местом (BANK)",      # Место появилось (раньше не было, теперь есть)
    "val_remove":   "Место пропало (BANK)",   # Место исчезло (было, теперь нет)
    "val_nochange": "Такое же место (BANK)",# Место не изменилось
    "val_up":       "Улучшил место (BANK)",       # Место улучшилось (меньше номер — выше место, например: с 5 на 2)
    "val_down":     "Хуже место (BANK)",     # Место ухудшилось (больше номер — ниже место, например: с 2 на 5)
    "val_norank":   "Нет места"             # Место отсутствовало и до, и после (None/NaN)
}

# --- Статусы сравнения для места (placeInRating) TB ---
STATUS_TB_PLACE = {
    "val_add":      "Новый с местом (TB)",        # Место появилось (раньше не было, теперь есть)
    "val_remove":   "Место пропало (TB)",     # Место исчезло (было, теперь нет)
    "val_nochange": "Такое же место (TB)",  # Место не изменилось
    "val_up":       "Улучшил место (TB)",         # Место улучшилось (меньше номер)
    "val_down":     "Хуже место (TB)",       # Место ухудшилось (больше номер)
    "val_norank":   "Нет места"             # Место отсутствовало и до, и после (None/NaN)
}

# --- Статусы сравнения для места (placeInRating) GOSB ---
STATUS_GOSB_PLACE = {
    "val_add":      "Новый с местом (GOSB)",      # Место появилось (раньше не было, теперь есть)
    "val_remove":   "Место пропало (GOSB)",   # Место исчезло (было, теперь нет)
    "val_nochange": "Такое же место (GOSB)",# Место не изменилось
    "val_up":       "Улучшил место (GOSB)",       # Место улучшилось (меньше номер)
    "val_down":     "Хуже место (GOSB)",     # Место ухудшилось (больше номер)
    "val_norank":   "Нет места"             # Место отсутствовало и до, и после (None/NaN)
}

CATEGORY_RANK_MAP = {
    "Вы в лидерах": 1,
    "Серебро": 2,
    "Бронза": 3,
    "Нужно поднажать": 4,
    "": 4,
    None: 4
}

category_compare_lookup = {
    (0, None, 0, None): {'desc_ru': 'Не участвовал', 'tag': 'NoShow'},
    (0, None, 1, 4): {'desc_ru': 'Новый участник без награды', 'tag': 'New'},
    (0, None, 1, 3): {'desc_ru': 'Новый призёр', 'tag': 'NewWin'},
    (0, None, 1, 2): {'desc_ru': 'Новый призёр', 'tag': 'NewWin'},
    (0, None, 1, 1): {'desc_ru': 'Новый призёр', 'tag': 'NewWin'},
    (1, 4, 0, None): {'desc_ru': 'Удалённый участник без награды', 'tag': 'Removed'},
    (1, 3, 0, None): {'desc_ru': 'Удалённый призёр', 'tag': 'RemovedWin'},
    (1, 2, 0, None): {'desc_ru': 'Удалённый призёр', 'tag': 'RemovedWin'},
    (1, 1, 0, None): {'desc_ru': 'Удалённый призёр', 'tag': 'RemovedWin'},
    (1, 4, 1, 4): {'desc_ru': 'Остался вне призеров', 'tag': 'NoChange'},
    (1, 4, 1, 3): {'desc_ru': 'Стал призёром', 'tag': 'Upgrade'},
    (1, 4, 1, 2): {'desc_ru': 'Стал призёром', 'tag': 'Upgrade'},
    (1, 4, 1, 1): {'desc_ru': 'Стал призёром', 'tag': 'Upgrade'},
    (1, 3, 1, 4): {'desc_ru': 'Лишился награды', 'tag': 'Lost'},
    (1, 3, 1, 3): {'desc_ru': 'Сохранил призовую позицию', 'tag': 'KeptSame'},
    (1, 3, 1, 2): {'desc_ru': 'Поднялся в рейтинге призеров', 'tag': 'WinUp'},
    (1, 3, 1, 1): {'desc_ru': 'Поднялся в рейтинге призеров', 'tag': 'WinUp'},
    (1, 2, 1, 4): {'desc_ru': 'Лишился награды', 'tag': 'Lost'},
    (1, 2, 1, 3): {'desc_ru': 'Снизил призовое место', 'tag': 'WinDown'},
    (1, 2, 1, 2): {'desc_ru': 'Сохранил призовую позицию', 'tag': 'KeptSame'},
    (1, 2, 1, 1): {'desc_ru': 'Поднялся в рейтинге призеров', 'tag': 'WinUp'},
    (1, 1, 1, 4): {'desc_ru': 'Лишился награды', 'tag': 'Lost'},
    (1, 1, 1, 3): {'desc_ru': 'Снизил призовое место', 'tag': 'WinDown'},
    (1, 1, 1, 2): {'desc_ru': 'Снизил призовое место', 'tag': 'WinDown'},
    (1, 1, 1, 1): {'desc_ru': 'Сохранил призовую позицию', 'tag': 'KeptSame'},
}

def setup_logger(log_dir, basename):
    """
    Создаёт логгер, который пишет и в файл (append, по дате), и в консоль.
    log_dir: путь к папке для логов
    basename: имя (без даты) для лог-файла
    """
    now = datetime.now()
    day_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H:%M:%S")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, f"{basename}_{day_str}.log")
    # Добавляем маркер начала новой сессии
    with open(log_path, "a", encoding="utf-8") as logf:
        logf.write(LOG_MESSAGES["LOGGER_SESSION_START"].format(date=day_str, time=time_str))
    # Стандартное подключение логгера
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    if logger.hasHandlers():
        logger.handlers.clear()
    fh = logging.FileHandler(log_path, encoding='utf-8', mode='a')
    fh.setLevel(LOG_LEVEL)
    ch = logging.StreamHandler()
    ch.setLevel(LOG_LEVEL)
    fmt = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s', "%Y-%m-%d %H:%M:%S")
    fh.setFormatter(fmt)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    logging.info(LOG_MESSAGES["LOGGER_ACTIVE_FILE"].format(path=log_path))
    return logger

def log_data_stats(df, label):
    if df.empty:
        logging.info(LOG_MESSAGES["DATAFRAME_EMPTY"].format(label=label))
        return
    n_rows = len(df)
    n_cols = len(df.columns)
    tournament_counts = df['tournamentId'].value_counts().to_dict()
    unique_tids = list(df['tournamentId'].unique())
    logging.info(LOG_MESSAGES["DATAFRAME_SHAPE"].format(label=label, n_rows=n_rows, n_cols=n_cols))
    logging.info(LOG_MESSAGES["TOURNAMENT_ID_LIST"].format(label=label, count=len(unique_tids), tids=unique_tids))
    for tid in unique_tids:
        count = tournament_counts.get(tid, 0)
        logging.debug(LOG_MESSAGES["TOURNAMENT_ID_PERSON_COUNT"].format(label=label, tid=tid, count=count))
    logging.info(LOG_MESSAGES["DATAFRAME_COLUMNS"].format(label=label, cols=list(df.columns)))


def log_compare_stats(compare_df):
    """Выводит сводную статистику по статусным колонкам таблицы сравнения."""
    n_rows = len(compare_df)
    logging.info(LOG_MESSAGES["COMPARE_TOTAL_ROWS"].format(n_rows=n_rows))
    for col in COMPARE_STATUS_COLUMNS:
        if col in compare_df.columns:
            counts = compare_df[col].value_counts(dropna=False).to_dict()
            logging.info(LOG_MESSAGES["COMPARE_COLUMN_COUNTS"].format(col=col, counts=counts))


def parse_float(val, context=None):
    """Преобразует значение в float, если возможно."""
    try:
        if val is None or (isinstance(val, str) and val.strip().lower() in {'', 'none', 'null'}):
            return None
        if isinstance(val, (int, float)):
            return round(float(val), 3)
        s = str(val)
        s = re.sub(r'[\s\u00A0\u2009]', '', s)
        s = re.sub(r"[^\d.,\-]", "", s)
        if s.count(',') > 0 and s.count('.') > 0:
            if s.rfind('.') > s.rfind(','):
                s = s.replace(',', '')
            else:
                s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '.')
        return round(float(s), 3)
    except Exception as ex:
        logging.error(
            LOG_MESSAGES["PARSE_FLOAT_ERROR"].format(val=val, ex=ex, context=context)
        )
        return None


def parse_int(val, context=None):
    """Преобразует значение в int, если возможно."""
    try:
        if val is None or (isinstance(val, str) and val.strip().lower() in {'', 'none', 'null'}):
            return None
        if isinstance(val, int):
            return val
        s = str(val)
        s = re.sub(r'[\s\u00A0\u2009]', '', s)
        s = re.sub(r"[^\d\-]", "", s)
        return int(s)
    except Exception as ex:
        logging.error(
            LOG_MESSAGES["PARSE_INT_ERROR"].format(val=val, ex=ex, context=context)
        )
        return None


def flatten_leader(leader, tournament_id, source_file):
    """Разворачивает запись лидера в плоскую структуру для DataFrame."""
    employee = leader.get('employeeNumber', 'N/A')
    context = f"файл={source_file}, турнир={tournament_id}, employee={employee}"
    # Лог: запуск обработки лидера
    logging.debug(LOG_MESSAGES["FLATTEN_LEADER_START"].format(
        employee=employee, tournament_id=tournament_id, source_file=source_file
    ))

    row = {
        'SourceFile': source_file,
        'tournamentId': tournament_id
    }
    for k, v in leader.items():
        if k in ("divisionRatings", "photoData"):
            continue
        if k in FLOAT_FIELDS:
            row[k] = parse_float(v, context)
        else:
            row[k] = v
    # divisionRatings
    if "divisionRatings" in leader and leader["divisionRatings"]:
        for div in leader["divisionRatings"]:
            group = div.get("groupCode")
            if not group:
                continue
            for field in ("groupId", "placeInRating", "ratingCategoryName"):
                colname = f"divisionRatings_{group}_{field}"
                if field in div:
                    value = div[field]
                    if colname in INT_FIELDS:
                        row[colname] = parse_int(value, context)
                    elif colname in FLOAT_FIELDS:
                        row[colname] = parse_float(value, context)
                    else:
                        row[colname] = value
    for f in FLOAT_FIELDS:
        if f not in row:
            row[f] = None
    for f in INT_FIELDS:
        if f not in row:
            row[f] = None
    return row


def process_json_file(filepath):
    """Загружает JSON-файл и превращает его в список словарей."""
    filename = os.path.basename(filepath)
    rows = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            js = json.load(f)
    except Exception as ex:
        logging.error(LOG_MESSAGES["PROCESS_JSON_LOAD_ERROR"].format(filepath=filepath, ex=ex))
        return []
    # Перебор турниров
    for tournament_key, records in js.items():
        entries = []
        if isinstance(records, list):
            entries = records
        elif isinstance(records, dict):
            entries = [records]
        else:
            logging.warning(LOG_MESSAGES["PROCESS_JSON_BAD_RECORD"].format(
                tournament_key=tournament_key, record=repr(records)[:100]))
            continue
        for record in entries:
            try:
                if not isinstance(record, dict):
                    logging.warning(LOG_MESSAGES["PROCESS_JSON_BAD_RECORD"].format(
                        tournament_key=tournament_key, record=repr(record)[:100]))
                    continue
                tournament = record.get("body", {}).get("tournament", {})
                tournament_id = tournament.get("tournamentId", tournament_key)
                leaders = tournament.get("leaders", [])
                if isinstance(leaders, dict):
                    leaders = list(leaders.values())
                elif not isinstance(leaders, list):
                    leaders = []
                if not leaders:
                    stub = {
                        'SourceFile': filename,
                        'tournamentId': tournament_id,
                        'employeeNumber': '00000000',
                        'lastName': 'None',
                        'firstName': 'None'
                    }
                    for field in FLOAT_FIELDS + INT_FIELDS:
                        stub[field] = None
                    rows.append(stub)
                    logging.info(LOG_MESSAGES["PROCESS_JSON_EMPTY_LEADERS"].format(
                        tournament_id=tournament_id, filename=filename))
                    continue
                for leader in leaders:
                    try:
                        row = flatten_leader(leader, tournament_id, filename)
                        rows.append(row)
                    except Exception as ex:
                        logging.error(LOG_MESSAGES["PROCESS_JSON_FLATTEN_LEADER_ERROR"].format(
                            filename=filename, tournament_id=tournament_id,
                            employee=leader.get('employeeNumber', 'N/A'), ex=ex))
            except Exception as ex:
                logging.error(LOG_MESSAGES["PROCESS_JSON_RECORD_ERROR"].format(
                    filename=filename, tournament_key=tournament_key, ex=ex))
    return rows

def filter_dataframe_by_tournaments(df, allowed_ids, filter_enabled, label="DataFrame"):
    """
    Фильтрует DataFrame по списку турниров в зависимости от параметров.
    
    Args:
        df: DataFrame для фильтрации
        allowed_ids: список разрешенных tournament_ids
        filter_enabled: флаг включения фильтрации
        label: метка для логирования
    
    Returns:
        Отфильтрованный DataFrame
    """
    original_count = len(df)
    
    # Если фильтрация отключена, возвращаем без изменений
    if not filter_enabled:
        logging.info(f"[{label}] Фильтрация турниров отключена, загружено {original_count} строк")
        return df
    
    # Если список турниров пустой, загружаем все
    if not allowed_ids:
        logging.info(f"[{label}] Список разрешенных турниров пустой, загружено {original_count} строк")
        return df
    
    # Применяем фильтрацию
    filtered_df = df[df['tournamentId'].isin(allowed_ids)]
    filtered_count = len(filtered_df)
    
    logging.info(f"[{label}] Применена фильтрация по турнирам: {original_count} -> {filtered_count} строк")
    logging.info(f"[{label}] Разрешенные турниры: {allowed_ids}")
    
    return filtered_df

def select_best_status_and_level(row, field_type="placeInRating"):
    """
    Выбирает лучший доступный статус из трех уровней (BANK -> TB -> GOSB).
    
    Args:
        row: строка DataFrame
        field_type: тип поля ("placeInRating" или "ratingCategoryName")
    
    Returns:
        tuple: (before_value, after_value, compare_status, level)
    """
    levels = ["BANK", "TB", "GOSB"]
    
    for level in levels:
        compare_col = f'divisionRatings_{level}_{field_type}_Compare'
        before_col = f'BEFORE_divisionRatings_{level}_{field_type}'
        after_col = f'AFTER_divisionRatings_{level}_{field_type}'
        
        if compare_col in row:
            status = row[compare_col]
            # Для placeInRating проверяем "Нет места", для ratingCategoryName - пустые значения
            if field_type == "placeInRating":
                if pd.notnull(status) and str(status).strip() != "Нет места":
                    return (
                        row.get(before_col), 
                        row.get(after_col), 
                        status, 
                        level
                    )
            else:  # ratingCategoryName
                if pd.notnull(status) and str(status).strip() not in ["", "Нет призового значения", "Не участвовал"]:
                    return (
                        row.get(before_col), 
                        row.get(after_col), 
                        status, 
                        level
                    )
    
    # Если не нашли подходящий статус, возвращаем значения из BANK уровня
    level = "BANK"
    compare_col = f'divisionRatings_{level}_{field_type}_Compare'
    before_col = f'BEFORE_divisionRatings_{level}_{field_type}'
    after_col = f'AFTER_divisionRatings_{level}_{field_type}'
    
    fallback_status = "Нет места" if field_type == "placeInRating" else "Нет призового значения"
    
    return (
        row.get(before_col), 
        row.get(after_col), 
        row.get(compare_col, fallback_status), 
        level
    )

def make_compare_sheet(df_before, df_after, sheet_name):
    logging.info(LOG_MESSAGES["COMPARE_SHEET_START"])

    join_keys = COMPARE_KEYS
    before_uniq = df_before.drop_duplicates(subset=join_keys, keep='last')
    after_uniq  = df_after.drop_duplicates(subset=join_keys, keep='last')
    all_keys = pd.concat([before_uniq[join_keys], after_uniq[join_keys]]).drop_duplicates()
    before_uniq = before_uniq.set_index(join_keys)
    after_uniq  = after_uniq.set_index(join_keys)
    before_uniq = before_uniq[COMPARE_FIELDS] if len(before_uniq) else pd.DataFrame(columns=COMPARE_FIELDS)
    after_uniq  = after_uniq[COMPARE_FIELDS]  if len(after_uniq) else pd.DataFrame(columns=COMPARE_FIELDS)
    before_uniq = before_uniq.add_prefix('BEFORE_')
    after_uniq  = after_uniq.add_prefix('AFTER_')
    compare_df = all_keys.set_index(join_keys) \
        .join(before_uniq, how='left') \
        .join(after_uniq, how='left') \
        .reset_index()

    # indicatorValue_Compare
    def value_compare(row):
        before = row.get('BEFORE_indicatorValue', None)
        after  = row.get('AFTER_indicatorValue', None)
        if pd.isnull(before) and not pd.isnull(after):
            return STATUS_INDICATOR['val_add']
        if not pd.isnull(before) and pd.isnull(after):
            return STATUS_INDICATOR['val_remove']
        if pd.isnull(before) and pd.isnull(after):
            return ""
        if before == after:
            return STATUS_INDICATOR['val_nochange']
        elif before > after:
            return STATUS_INDICATOR['val_down']
        else:
            return STATUS_INDICATOR['val_up']
    compare_df['indicatorValue_Compare'] = compare_df.apply(value_compare, axis=1)

    def rang_compare(row, before_col, after_col, status_dict):
        before = row.get(f'BEFORE_{before_col}', None)
        after  = row.get(f'AFTER_{after_col}', None)
        if pd.isnull(before) and not pd.isnull(after):
            return status_dict['val_add']
        if not pd.isnull(before) and pd.isnull(after):
            return status_dict['val_remove']
        if pd.isnull(before) and pd.isnull(after):
            return status_dict.get('val_norank', 'Нет места')
        if before == after:
            return status_dict['val_nochange']
        elif before > after:
            return status_dict['val_up']
        else:
            return status_dict['val_down']

    compare_df['divisionRatings_BANK_placeInRating_Compare'] = compare_df.apply(
        lambda row: rang_compare(row, 'divisionRatings_BANK_placeInRating', 'divisionRatings_BANK_placeInRating', STATUS_BANK_PLACE), axis=1)
    compare_df['divisionRatings_TB_placeInRating_Compare'] = compare_df.apply(
        lambda row: rang_compare(row, 'divisionRatings_TB_placeInRating', 'divisionRatings_TB_placeInRating', STATUS_TB_PLACE), axis=1)
    compare_df['divisionRatings_GOSB_placeInRating_Compare'] = compare_df.apply(
        lambda row: rang_compare(row, 'divisionRatings_GOSB_placeInRating', 'divisionRatings_GOSB_placeInRating', STATUS_GOSB_PLACE), axis=1)

    # === Функция сравнения категорий ===
    def category_compare_enhanced(row, colname):
        before_cat = row.get(f'BEFORE_{colname}')
        after_cat = row.get(f'AFTER_{colname}')
        before_present = 0 if pd.isnull(before_cat) or before_cat == "" else 1
        after_present = 0 if pd.isnull(after_cat) or after_cat == "" else 1
        b_cat = CATEGORY_RANK_MAP.get(before_cat, None) if before_present else None
        a_cat = CATEGORY_RANK_MAP.get(after_cat, None) if after_present else None
        key = (before_present, b_cat, after_present, a_cat)
        result = category_compare_lookup.get(key)
        return result["desc_ru"] if result else ""

    for group, col in [
        ('BANK',   'divisionRatings_BANK_ratingCategoryName'),
        ('TB',     'divisionRatings_TB_ratingCategoryName'),
        ('GOSB',   'divisionRatings_GOSB_ratingCategoryName'),
    ]:
        compare_df[f"{col}_Compare"] = compare_df.apply(lambda row: category_compare_enhanced(row, col), axis=1)

    # === Объединение статусов по приоритету BANK -> TB -> GOSB ===
    logging.info("Формирование объединенных статусов по лучшему доступному уровню...")
    
    # Для placeInRating
    place_results = compare_df.apply(lambda row: select_best_status_and_level(row, "placeInRating"), axis=1)
    compare_df['BEFORE_placeInRating_Best'] = [res[0] for res in place_results]
    compare_df['AFTER_placeInRating_Best'] = [res[1] for res in place_results]
    compare_df['placeInRating_Compare_Best'] = [res[2] for res in place_results]
    compare_df['placeInRating_Level'] = [res[3] for res in place_results]
    
    # Для ratingCategoryName
    category_results = compare_df.apply(lambda row: select_best_status_and_level(row, "ratingCategoryName"), axis=1)
    compare_df['BEFORE_ratingCategoryName_Best'] = [res[0] for res in category_results]
    compare_df['AFTER_ratingCategoryName_Best'] = [res[1] for res in category_results]
    compare_df['ratingCategoryName_Compare_Best'] = [res[2] for res in category_results]
    compare_df['ratingCategoryName_Level'] = [res[3] for res in category_results]
    
    logging.info("Объединенные статусы сформированы")

    final_cols = COMPARE_KEYS + COMPARE_STATUS_COLUMNS + ['BEFORE_' + c for c in COMPARE_FIELDS] + ['AFTER_' + c for c in COMPARE_FIELDS]
    # Добавляем новые колонки в final_cols
    final_cols += [
        'BEFORE_placeInRating_Best', 'AFTER_placeInRating_Best', 'placeInRating_Compare_Best', 'placeInRating_Level',
        'BEFORE_ratingCategoryName_Best', 'AFTER_ratingCategoryName_Best', 'ratingCategoryName_Compare_Best', 'ratingCategoryName_Level'
    ]
    compare_df = compare_df.reindex(columns=final_cols)

    # --- Фильтрация по tournamentId ---
    if ALLOWED_TOURNAMENT_IDS:
        compare_df = compare_df[compare_df['tournamentId'].isin(ALLOWED_TOURNAMENT_IDS)]
        logging.info(LOG_MESSAGES["COMPARE_SHEET_FILTERED"].format(count=len(compare_df)))

    # --- Фильтрация строк без изменений ---
    status_cols = COMPARE_STATUS_COLUMNS

    def is_any_change(row):
        for col in status_cols:
            val = str(row.get(col, "")).strip()
            if val not in NOCHANGE_STATUSES:
                return True
        return False

    compare_df = compare_df[compare_df.apply(is_any_change, axis=1)].reset_index(drop=True)
    logging.info(LOG_MESSAGES["COMPARE_SHEET_FINAL"].format(count=len(compare_df)))
    return compare_df, sheet_name


def add_smart_table(writer, df, sheet_name, table_name, freeze_map=None):
    logging.info(LOG_MESSAGES["SMART_TABLE_EXPORT_START"].format(sheet=sheet_name))
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    try:
        worksheet = writer.sheets[sheet_name]

        # Жирный заголовок и перенос строк
        for i, cell in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1)), 1):
            cell.font = Font(bold=True)
            col_name = str(cell.value or "")
            # Для турниров, статусов, длинных — перенос по словам
            if ("турнир" in col_name.lower()) or ("place" in col_name.lower()) or len(col_name) > 14:
                cell.alignment = Alignment(wrap_text=True)

        # Автоширина столбцов — по размеру значения, не меньше 10, не больше 30
        for i, column in enumerate(df.columns, 1):
            col_name = str(column)
            values = df[column].astype(str) if not df.empty else []
            max_length = max([len(str(col_name))] + [len(x) for x in values])
            width = min(max(max_length + 2, 10), 30)
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Автофильтр
        worksheet.auto_filter.ref = worksheet.dimensions

        # Freeze panes
        if freeze_map:
            key = sheet_name.split()[0].upper()
            if key in freeze_map:
                worksheet.freeze_panes = freeze_map[key]
        logging.info(LOG_MESSAGES["SMART_TABLE_FORMATTED"].format(sheet=sheet_name))
    except Exception as ex:
        logging.error(LOG_MESSAGES["SMART_TABLE_ERROR"].format(sheet=sheet_name, err=ex))
        logging.error(traceback.format_exc())



def apply_status_colors(writer, df, sheet_name, status_color_map, status_columns):
    logging.info(LOG_MESSAGES["STATUS_COLOR_START"].format(sheet=sheet_name))
    worksheet = writer.sheets[sheet_name]
    for col_name in status_columns:
        if col_name not in df.columns:
            continue
        col_idx = df.columns.get_loc(col_name) + 1
        for row_idx, value in enumerate(df[col_name], 2):
            status = str(value)
            color_info = status_color_map.get(status, {})
            fill_color = color_info.get("fill", None)
            font_color = color_info.get("font", None)
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if fill_color:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_color.lstrip('#'))
            if font_color:
                cell.font = Font(color=font_color.lstrip('#'))
    logging.info(LOG_MESSAGES["STATUS_COLOR_DONE"].format(sheet=sheet_name))

def add_status_legend(writer, legend_data, sheet_name=SHEET_NAMES['status_legend']):
    # legend_data: массив из STATUS_LEGEND_FULL
    columns = ["Статус", "Код", "Описание", "Цвет заливки", "Цвет текста", "Где применяется", "Группа"]
    df_legend = pd.DataFrame(legend_data, columns=columns)
    df_legend.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]

    # Цветовые заливки + автоширина
    for row_idx, row in enumerate(legend_data, start=2):
        fill = row[3].lstrip('#')
        font = row[4].lstrip('#')
        ws.cell(row=row_idx, column=4).fill = PatternFill(fill_type="solid", fgColor=fill)
        ws.cell(row=row_idx, column=4).font = Font(color=font)
        ws.cell(row=row_idx, column=5).font = Font(color=font)
    # Жирный заголовок
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Автоширина
    from openpyxl.utils import get_column_letter
    for i, column_cells in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    logging.info(LOG_MESSAGES["STATUS_LEGEND_ADD"].format(sheet=sheet_name))



def build_final_sheet_fast(compare_df, allowed_ids, out_prefix, category_rank_map, df_before, df_after, log, sheet_name="FINAL"):
    """Строит итоговый лист по всем турнирам и сотрудникам. Оптимизированная версия: lookup-структуры вместо фильтрации."""
    log.info(LOG_MESSAGES["FINAL_BUILD_START"])
    if allowed_ids:
        tournaments = list(allowed_ids)
    else:
        tournaments = sorted(compare_df['tournamentId'].dropna().unique())

    emp_cols = ['employeeNumber', 'lastName', 'firstName']
    employees = compare_df[emp_cols].drop_duplicates().sort_values(emp_cols)
    log.info(LOG_MESSAGES["FINAL_UNIQUE_EMPLOYEES"].format(num_employees=len(employees)))
    total_loops = len(employees) * len(tournaments)
    log.info(LOG_MESSAGES["FINAL_TOTAL_LOOPS"].format(loops=total_loops))

    # --- Быстрый доступ по сотруднику+турнир (tuple index)
    indexed = compare_df.set_index(['employeeNumber', 'lastName', 'firstName', 'tournamentId'])
    # --- Быстрый доступ к признаку "был ли в before/after"
    before_pairs = set(zip(df_before['employeeNumber'], df_before['tournamentId']))
    after_pairs = set(zip(df_after['employeeNumber'], df_after['tournamentId']))

    # Теперь используем объединенный статус категорий
    best_status_col = 'ratingCategoryName_Compare_Best'

    result_rows = []
    status_counter = {t_id: {} for t_id in tournaments}

    for emp_idx, (_, emp) in enumerate(employees.iterrows(), 1):
        emp_key = (emp['employeeNumber'], emp['lastName'], emp['firstName'])
        row = {col: emp[col] for col in emp_cols}

        for t_id in tournaments:
            idx = emp_key + (t_id,)
            best_val = None

            subset = indexed.loc[idx] if idx in indexed.index else None
            if subset is not None:
                # Используем объединенный статус, который уже выбран по приоритету BANK->TB->GOSB
                val = subset.get(best_status_col)
                if pd.notnull(val) and str(val).strip() not in ['', 'Нет призового значения', 'Не участвовал']:
                    best_val = val

            was_in_before = (emp['employeeNumber'], t_id) in before_pairs
            was_in_after = (emp['employeeNumber'], t_id) in after_pairs

            if best_val is not None:
                final_value = best_val
            elif was_in_before or was_in_after:
                final_value = "Нет призового значения"
            else:
                final_value = "Не участвовал"

            status_counter[t_id][final_value] = status_counter[t_id].get(final_value, 0) + 1

            row[t_id] = final_value
        result_rows.append(row)

    final_df = pd.DataFrame(result_rows)
    log.info(LOG_MESSAGES["FINAL_TABLE_DONE"].format(shape=f"{final_df.shape[0]} x {final_df.shape[1]}"))
    # Подробное логирование по каждому турниру
    for t_id in tournaments:
        log.debug(LOG_MESSAGES["FINAL_TOURN_STATUS"].format(sheet=sheet_name, tid=t_id))
        for status, count in status_counter[t_id].items():
            log.debug(LOG_MESSAGES["FINAL_TOURN_STATUS_ROW"].format(sheet=sheet_name, status=status, count=count))
    return final_df, tournaments

def build_tournament_fullname_map(catalog_dir, schedule_csv, contest_csv):
    """Создает маппинг {tournamentId: <FULL_NAME> ("<tournamentId>")} для всех турниров."""
    try:
        schedule_path = os.path.join(catalog_dir, schedule_csv)
        contest_path = os.path.join(catalog_dir, contest_csv)
        df_schedule = pd.read_csv(schedule_path, sep=";", dtype=str)
        df_contest = pd.read_csv(contest_path, sep=";", dtype=str)
    except Exception as ex:
        logging.error(f"[build_tournament_fullname_map] Ошибка чтения CSV: {ex}")
        return {}

    # 1. Соединяем schedule и contest по CONTEST_CODE
    df_merged = pd.merge(df_schedule, df_contest, how="left", on="CONTEST_CODE")
    # 2. Для каждого TOURNAMENT_CODE формируем отображение
    tid_to_full = {}
    for idx, row in df_merged.iterrows():
        tid = row["TOURNAMENT_CODE"]
        full_name = row.get("FULL_NAME", "").strip()
        if not full_name:
            continue
        tid_to_full[tid] = f'{full_name} ("{tid}")'
    return tid_to_full

def replace_tournamentid_with_fullname(df, tid_map):
    """Переименовывает все колонки с tournamentId и значения по маппингу."""
    # Переименование колонок-турниров в финальных таблицах (FINAL, FINAL_PLACE)
    col_rename = {tid: tid_map.get(tid, tid) for tid in df.columns if tid in tid_map}
    df = df.rename(columns=col_rename)

    # Для compare_df, final_df и других: если есть колонка tournamentId — заменить значения
    if "tournamentId" in df.columns:
        df["tournamentId"] = df["tournamentId"].apply(lambda x: tid_map.get(x, x) if pd.notnull(x) else x)
    return df


def format_compare_dataframe(compare_df, export_columns):
    """
    Форматирует итоговый DataFrame для COMPARE:
    - объединяет соответствующие колонки BEFORE/AFTER (SourceFile, terDivisionName, groupId)
    - переставляет порядок колонок согласно export_columns
    - заменяет 'tournamentId' на 'tournamentName' только в порядке колонок
    """
    compare_df = compare_df.copy()

    # 1. Объединённые колонки
    compare_df['SourceFile'] = compare_df['BEFORE_SourceFile'].fillna('') + " => " + compare_df['AFTER_SourceFile'].fillna('')
    compare_df['terDivisionName'] = compare_df['BEFORE_terDivisionName'].fillna('') + " => " + compare_df['AFTER_terDivisionName'].fillna('')
    compare_df['divisionRatings_TB_groupId'] = compare_df['BEFORE_divisionRatings_TB_groupId'].astype(str).replace('nan', '') + " => " + compare_df['AFTER_divisionRatings_TB_groupId'].astype(str).replace('nan', '')
    compare_df['divisionRatings_GOSB_groupId'] = compare_df['BEFORE_divisionRatings_GOSB_groupId'].astype(str).replace('nan', '') + " => " + compare_df['AFTER_divisionRatings_GOSB_groupId'].astype(str).replace('nan', '')

    # Порядок колонок: заменяем 'tournamentId' на 'tournamentName', но саму колонку не удаляем
    columns_present = []
    for col in export_columns:
        if col == 'tournamentId' and 'tournamentName' in compare_df.columns:
            columns_present.append('tournamentName')
        elif col in compare_df.columns:
            columns_present.append(col)
    extra_cols = [col for col in compare_df.columns if col not in columns_present]
    return compare_df[columns_present + extra_cols]



def build_final_place_sheet_from_compare(compare_df, allowed_ids, df_before, df_after, log, sheet_name="FINAL_PLACE"):
    """
    Строит сводную таблицу по статусам placeInRating_Compare (BANK > TB > GOSB > Не участвовал).
    Подсчёт статусов идёт только по одному выбранному для турнира уровню, без дублей.
    """
    log.info(LOG_MESSAGES["PLACE_BUILD_START"].format(sheet=sheet_name))
    if allowed_ids:
        tournaments = list(allowed_ids)
    else:
        tournaments = sorted(compare_df['tournamentId'].dropna().unique())

    emp_cols = ['employeeNumber', 'lastName', 'firstName']
    employees = compare_df[emp_cols].drop_duplicates().sort_values(emp_cols)
    log.info(LOG_MESSAGES["PLACE_UNIQUE_EMPLOYEES"].format(sheet=sheet_name, num_employees=len(employees)))
    log.info(LOG_MESSAGES["PLACE_TOURNAMENTS"].format(sheet=sheet_name, num_tournaments=len(tournaments)))

    # Теперь используем объединенный статус места
    best_place_col = 'placeInRating_Compare_Best'

    # Быстрые индексы
    indexed = compare_df.set_index(['employeeNumber', 'lastName', 'firstName', 'tournamentId'])
    before_pairs = set(zip(df_before['employeeNumber'], df_before['tournamentId']))
    after_pairs = set(zip(df_after['employeeNumber'], df_after['tournamentId']))

    result_rows = []
    status_counter = {t_id: {} for t_id in tournaments}

    # Строим по объединенному статусу мест
    for _, emp in employees.iterrows():
        emp_key = (emp['employeeNumber'], emp['lastName'], emp['firstName'])
        row = {col: emp[col] for col in emp_cols}

        for t_id in tournaments:
            idx = emp_key + (t_id,)
            value = None
            
            if idx in indexed.index:
                rec = indexed.loc[idx]
                # Используем объединенный статус, который уже выбран по приоритету BANK->TB->GOSB
                v = rec.get(best_place_col, None)
                if pd.notnull(v) and str(v).strip() not in ['', 'Нет места', 'Нет значения ранга', 'Не участвовал']:
                    value = v

            was_in_before = (emp['employeeNumber'], t_id) in before_pairs
            was_in_after = (emp['employeeNumber'], t_id) in after_pairs

            if value is not None:
                final_value = value
            elif was_in_before or was_in_after:
                final_value = "Нет значения ранга"
            else:
                final_value = "Не участвовал"

            row[t_id] = final_value
            status_counter[t_id][final_value] = status_counter[t_id].get(final_value, 0) + 1

        result_rows.append(row)

    final_place_df = pd.DataFrame(result_rows)
    log.info(LOG_MESSAGES["PLACE_TABLE_DONE"].format(sheet=sheet_name, shape=f"{final_place_df.shape[0]} x {final_place_df.shape[1]}"))

    # Логирование по турнирам
    for t_id in tournaments:
        log.debug(LOG_MESSAGES["PLACE_TOURN_STATUS"].format(sheet=sheet_name, tid=t_id))
        for status, count in status_counter[t_id].items():
            log.debug(LOG_MESSAGES["PLACE_TOURN_STATUS_ROW"].format(sheet=sheet_name, status=status, count=count))

    # Дополнительно: вернуть структуру tournament_level если потребуется для диагностики
    return final_place_df, tournaments

def apply_stat_grp_conditional_formatting(writer, sheet_name, stat_prefixes=('stat_', 'grp_'), log=None):
    ws = writer.sheets[sheet_name]
    from openpyxl.styles import PatternFill

    if log:
        log.info(LOG_MESSAGES["COND_FMT_START"].format(sheet=sheet_name, prefixes=stat_prefixes))

    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        col_letter = col[0].column_letter
        header = ws[f"{col_letter}1"].value
        if not header or not any(header.startswith(prefix) for prefix in stat_prefixes):
            continue
        col_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        # 1. Правило: если 0 — белый
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['0'],
                       stopIfTrue=True,
                       font=Font(color="E0E0E0"))
        )
        # 2. Цветовая шкала для >0
        ws.conditional_formatting.add(
            col_range,
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B',
            )
        )
        if log:
            log.debug(LOG_MESSAGES["COND_FMT_COL"].format(sheet=sheet_name, col=header))
    if log:
        log.info(LOG_MESSAGES["COND_FMT_FINISH"].format(sheet=sheet_name, prefixes=stat_prefixes))

def add_status_count_and_top3(df, status_cols, all_statuses, log, is_final_place=False):
    """
    Добавляет к DataFrame счетчики по статусам, top-3 (названия), и (для FINAL) — группы.
    """
    exclude = {"Не участвовал", "Нет призового значения", "Остался вне призеров"}
    stat_names = [s for s in all_statuses if s not in exclude]
    group_names = [g[0] for g in STATUS_GROUPS] if not is_final_place else []

    stat_cols = [f"stat_{s}" for s in stat_names]
    group_cols = [f"grp_{g}" for g in group_names]
    new_columns = list(df.columns) + stat_cols + ['TOP1', 'TOP2', 'TOP3']
    if not is_final_place:
        new_columns += group_cols + ['GRP_MAX']

    result_rows = []
    for idx, row in df.iterrows():
        stats = {s: 0 for s in stat_names}
        vals = [row.get(c, None) for c in status_cols]
        for v in vals:
            if v in stats:
                stats[v] += 1
        # TOP-3
        stat_items = sorted(stats.items(), key=lambda x: (-x[1], stat_names.index(x[0])))
        used_names = set()
        tops = []
        items_left = stat_items.copy()
        for _ in range(3):
            if not items_left or items_left[0][1] <= 0:
                tops.append('-')
                continue
            maxval = items_left[0][1]
            names = [n for n, cnt in items_left if cnt == maxval]
            tops.append(', '.join(names))
            used_names.update(names)
            items_left = [(n, cnt if n not in names else -1) for n, cnt in items_left]
            items_left = sorted(items_left, key=lambda x: (-x[1], stat_names.index(x[0])))

        # Подсчет по группам
        group_counts = []
        grp_max = "-"
        if not is_final_place:
            max_val = 0
            for gname, gstatuses in STATUS_GROUPS:
                cnt = sum(stats.get(s, 0) for s in gstatuses)
                group_counts.append(cnt)
                if cnt > max_val:
                    max_val = cnt
            max_groups = [group_names[i] for i, val in enumerate(group_counts) if val == max_val and val > 0]
            if max_groups:
                # Выбираем только первую по приоритету (обычно это Группа 1, если она есть)
                best_group = max_groups[0]
                grp_max = f"({best_group}) {GROUP_DESC_DICT[best_group]}"
            else:
                grp_max = "-"

        new_row = list(row.values)
        new_row += [stats[s] for s in stat_names]
        new_row += tops
        if not is_final_place:
            new_row += group_counts
            new_row += [grp_max]
        result_rows.append(new_row)
    log.info(LOG_MESSAGES["ADD_STATUSES_SUMMARY"].format(
        columns=stat_cols + ['TOP1', 'TOP2', 'TOP3'] + (group_cols + ['GRP_MAX'] if not is_final_place else [])
    ))
    return pd.DataFrame(result_rows, columns=new_columns), stat_names, group_cols


def export_and_log(writer, df, sheet_name, log, freeze_map=None):
    """
    Экспортирует DataFrame в Excel и логирует событие экспорта.
    Возвращает текст лога (для дальнейшего использования или вывода).
    """
    add_smart_table(writer, df, sheet_name, "SMART_" + sheet_name, freeze_map=freeze_map)
    msg = LOG_MESSAGES["EXPORT_SHEET"].format(sheet=sheet_name, rows=df.shape[0])
    log.info(msg)
    return msg

def get_status_distribution(df, status_list, colnames):
    """Возвращает словарь {status: count} по списку колонок с турнирами."""
    dist = {status: 0 for status in status_list}
    for col in colnames:
        if col not in df.columns:
            continue
        vals = df[col].value_counts()
        for s in dist.keys():
            dist[s] += int(vals.get(s, 0))
    return dist

def get_group_distribution(df, group_cols):
    """Возвращает распределение по группам."""
    result = {}
    for col in group_cols:
        if col in df.columns:
            result[col] = int(df[col].sum())
    return result

def main():
    """Основная точка входа в программу."""
    logger = setup_logger(LOG_DIR, LOG_BASENAME)

    # === Загрузка справочников и подготовка соответствия TournamentID → FULL_NAME ===
    tid_to_fullname = build_tournament_fullname_map(
        CATALOG_DIR, TOURNAMENT_SCHEDULE_CSV, CONTEST_DATA_CSV
    )
    logger.info(LOG_MESSAGES["MAIN_TOURNAMENT_DESCRIPTIONS_LOADED"].format(count=len(tid_to_fullname)))

    t_start = datetime.now()
    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S")

    # --- Загрузка данных ---
    logger.info(LOG_MESSAGES["MAIN_BEFORE_READ"].format(sheet=SHEET_NAMES['before'], path=os.path.join(SOURCE_DIR, BEFORE_FILENAME)))
    t_beg_before = datetime.now()
    rows_before = process_json_file(os.path.join(SOURCE_DIR, BEFORE_FILENAME))
    df_before = pd.DataFrame(rows_before)
    df_before['tournamentName'] = df_before['tournamentId'].map(tid_to_fullname)
    t_end_before = datetime.now()
    logger.info(LOG_MESSAGES["MAIN_BEFORE_LOADED"].format(count=len(df_before), sheet=SHEET_NAMES['before']))
    log_data_stats(df_before, SHEET_NAMES['before'])

    logger.info(LOG_MESSAGES["MAIN_AFTER_READ"].format(sheet=SHEET_NAMES['after'], path=os.path.join(SOURCE_DIR, AFTER_FILENAME)))
    t_beg_after = datetime.now()
    rows_after = process_json_file(os.path.join(SOURCE_DIR, AFTER_FILENAME))
    df_after = pd.DataFrame(rows_after)
    df_after['tournamentName'] = df_after['tournamentId'].map(tid_to_fullname)
    t_end_after = datetime.now()
    logger.info(LOG_MESSAGES["MAIN_AFTER_LOADED"].format(count=len(df_after), sheet=SHEET_NAMES['after']))
    log_data_stats(df_after, SHEET_NAMES['after'])

    # --- Фильтрация турниров для листов BEFORE и AFTER (при необходимости) ---
    df_before = filter_dataframe_by_tournaments(
        df_before, ALLOWED_TOURNAMENT_IDS, FILTER_TOURNAMENTS_IN_BEFORE_AFTER, SHEET_NAMES['before']
    )
    df_after = filter_dataframe_by_tournaments(
        df_after, ALLOWED_TOURNAMENT_IDS, FILTER_TOURNAMENTS_IN_BEFORE_AFTER, SHEET_NAMES['after']
    )

    # --- Анализ турниров ---
    before_tids = set(df_before['tournamentId'].unique())
    after_tids = set(df_after['tournamentId'].unique())
    added_tids = after_tids - before_tids
    removed_tids = before_tids - after_tids
    common_tids = before_tids & after_tids

    logger.info(LOG_MESSAGES["MAIN_TOURNAMENTS_INFO"].format(
        sheet_before=SHEET_NAMES['before'], before_count=len(before_tids),
        sheet_after=SHEET_NAMES['after'], after_count=len(after_tids)
    ))
    logger.info(LOG_MESSAGES["MAIN_TOURNAMENTS_NEW"].format(
        sheet_after=SHEET_NAMES['after'], count=len(added_tids), ids=list(added_tids)))
    logger.info(LOG_MESSAGES["MAIN_TOURNAMENTS_REMOVED"].format(
        sheet_before=SHEET_NAMES['before'], count=len(removed_tids), ids=list(removed_tids)))
    logger.info(LOG_MESSAGES["MAIN_TOURNAMENTS_COMMON"].format(
        count=len(common_tids), ids=list(common_tids)))

    # --- Приведение колонок к общему виду ---
    all_cols = PRIORITY_COLS.copy()
    all_cols += [c for c in set(df_before.columns).union(df_after.columns) if c not in all_cols]
    df_before = df_before.reindex(columns=all_cols)
    df_after = df_after.reindex(columns=all_cols)
    logger.info(LOG_MESSAGES["MAIN_COLUMNS_ALIGNED"].format(
        sheet_before=SHEET_NAMES['before'], sheet_after=SHEET_NAMES['after']))

    # --- Формируем COMPARE ---
    t_beg_compare = datetime.now()
    compare_df, sheet_compare = make_compare_sheet(df_before, df_after, SHEET_NAMES['compare'])
    compare_df['tournamentName'] = compare_df['tournamentId'].map(tid_to_fullname)
    compare_df = format_compare_dataframe(compare_df, COMPARE_EXPORT_COLUMNS)
    t_end_compare = datetime.now()
    logger.info(LOG_MESSAGES["MAIN_COMPARE_DONE"].format(
        sheet=sheet_compare, count=len(compare_df)))
    log_compare_stats(compare_df)

    # --- Финальная таблица (FINAL) ---
    t_beg_final = datetime.now()
    final_df, tournaments = build_final_sheet_fast(
        compare_df, ALLOWED_TOURNAMENT_IDS, "FINAL_", CATEGORY_RANK_MAP, df_before, df_after, logger, sheet_name=SHEET_NAMES['final']
    )
    logger.info(LOG_MESSAGES["MAIN_FINAL_DONE"].format(
        sheet=SHEET_NAMES['final'], shape=final_df.shape))

    # Финальная таблица по place (FINAL_PLACE)
    final_place_df, tournaments_place = build_final_place_sheet_from_compare(
        compare_df, ALLOWED_TOURNAMENT_IDS, df_before, df_after, logger, sheet_name=SHEET_NAMES['final_place']
    )
    logger.info(LOG_MESSAGES["MAIN_FINAL_PLACE_DONE"].format(
        sheet=SHEET_NAMES['final_place'], shape=final_place_df.shape))
    t_end_final = datetime.now()

    # --- Подсчет TOP-3 и групп (группы только для FINAL) ---
    final_df_stat, final_status_names, final_group_cols = add_status_count_and_top3(
        final_df, tournaments, FINAL_STATUS_LIST, logger, is_final_place=False
    )
    logger.info(LOG_MESSAGES["MAIN_FINAL_TOP3"].format(sheet=SHEET_NAMES['final']))

    final_place_df_stat, final_place_status_names, _ = add_status_count_and_top3(
        final_place_df, tournaments_place, FINAL_PLACE_STATUS_LIST, logger, is_final_place=True
    )
    logger.info(LOG_MESSAGES["MAIN_FINAL_PLACE_TOP3"].format(sheet=SHEET_NAMES['final_place']))

    # --- Сборка статистики по статусам и группам ---
    final_status_dist = get_status_distribution(final_df_stat, FINAL_STATUS_LIST, tournaments)
    final_place_status_dist = get_status_distribution(final_place_df_stat, FINAL_PLACE_STATUS_LIST, tournaments_place)
    final_groups_dist = get_group_distribution(final_df_stat, final_group_cols) if final_group_cols else {}

    # --- Экспорт в Excel ---
    base, ext = os.path.splitext(RESULT_EXCEL)
    result_excel_ts = f"{base}_{ts}{ext}"
    out_excel = os.path.join(TARGET_DIR, result_excel_ts)

    def export_final_sheet_with_names(writer, df, tournaments, sheet_name):
        df_export = df.copy()
        tournament_cols = [c for c in df.columns if c in tournaments]
        rename_map = {tid: tid_to_fullname.get(tid, tid) for tid in tournament_cols}
        df_export = df_export.rename(columns=rename_map)
        add_smart_table(writer, df_export, sheet_name, "SMART_" + sheet_name, freeze_map=freeze_map)
        ws = writer.sheets[sheet_name]
        for i, col in enumerate(df_export.columns, 1):
            col_name = str(col)
            cell = ws.cell(row=1, column=i)
            if col_name in rename_map.values():
                cell.alignment = Alignment(wrap_text=True)
            values = df_export[col].astype(str) if not df_export.empty else []
            max_length = max([len(str(col_name))] + [len(x) for x in values])
            width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[get_column_letter(i)].width = width

    t_beg_export = datetime.now()
    with pd.ExcelWriter(out_excel, engine='openpyxl') as writer:
        export_and_log(writer, df_before, SHEET_NAMES['before'], logger, freeze_map)
        export_and_log(writer, df_after, SHEET_NAMES['after'], logger, freeze_map)

        # Для COMPARE: удаляем tournamentId только при экспорте
        compare_export_df = compare_df.copy()
        if 'tournamentId' in compare_export_df.columns:
            compare_export_df = compare_export_df.drop(columns=['tournamentId'])
        export_and_log(writer, compare_export_df, SHEET_NAMES['compare'], logger, freeze_map)

        # Финальные таблицы с заголовками-названиями турниров
        export_final_sheet_with_names(writer, final_df_stat, tournaments, SHEET_NAMES['final'])
        export_final_sheet_with_names(writer, final_place_df_stat, tournaments_place, SHEET_NAMES['final_place'])

        apply_stat_grp_conditional_formatting(writer, SHEET_NAMES['final'], ('stat_', 'grp_'), log=logger)
        logger.info(LOG_MESSAGES["MAIN_STAT_COND_FMT"].format(sheet=SHEET_NAMES['final']))
        apply_stat_grp_conditional_formatting(writer, SHEET_NAMES['final_place'], ('stat_', 'grp_'), log=logger)
        logger.info(LOG_MESSAGES["MAIN_STAT_COND_FMT"].format(sheet=SHEET_NAMES['final_place']))

        # Цветовая раскраска
        apply_status_colors(writer, final_df_stat, SHEET_NAMES['final'], STATUS_COLORS_DICT, tournaments + final_status_names + ['TOP1', 'TOP2', 'TOP3'])
        logger.info(LOG_MESSAGES["MAIN_COLORS_APPLIED"].format(sheet=SHEET_NAMES['final']))
        apply_status_colors(writer, final_place_df_stat, SHEET_NAMES['final_place'], STATUS_COLORS_DICT, tournaments_place + final_place_status_names + ['TOP1', 'TOP2', 'TOP3'])
        logger.info(LOG_MESSAGES["MAIN_COLORS_APPLIED"].format(sheet=SHEET_NAMES['final_place']))
        apply_status_colors(writer, compare_export_df, SHEET_NAMES['compare'], STATUS_COLORS_DICT, COMPARE_STATUS_COLUMNS)
        logger.info(LOG_MESSAGES["MAIN_COLORS_APPLIED"].format(sheet=SHEET_NAMES['compare']))

        add_status_legend(writer, STATUS_LEGEND_FULL, sheet_name=SHEET_NAMES['status_legend'])
        logger.info(LOG_MESSAGES["MAIN_LEGEND_ADDED"])

        try:
            workbook = writer.book
            if SHEET_NAMES['final'] in workbook.sheetnames:
                workbook.active = workbook.sheetnames.index(SHEET_NAMES['final'])
        except Exception as ex:
            logger.warning(LOG_MESSAGES["MAIN_FINAL_SET_ACTIVE_SHEET_FAIL"].format(sheet=SHEET_NAMES['final'], ex=ex))

        logger.info(LOG_MESSAGES["MAIN_EXCEL_EXPORT"].format(path=out_excel))
    t_end_export = datetime.now()

    # --- Сводка по времени и группам/статусам ---
    summary = SUMMARY_TEMPLATE_EXT.format(
        tourn=len(tournaments),
        emps=len(final_df),
        changes=len(compare_df),
        t1=(t_end_before - t_beg_before).total_seconds(),
        t2=(t_end_after - t_beg_after).total_seconds(),
        t3=(t_end_compare - t_beg_compare).total_seconds(),
        t4=(t_end_final - t_beg_final).total_seconds(),
        t5=(t_end_export - t_beg_export).total_seconds(),
        tt=(t_end_export - t_start).total_seconds(),
        final_statuses=final_status_dist,
        final_place_statuses=final_place_status_dist,
        final_groups=final_groups_dist,
    )
    logger.info(summary)


if __name__ == "__main__":
    main()